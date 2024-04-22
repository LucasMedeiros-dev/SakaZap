import sys
from PySide6 import QtWidgets, QtCore, QtGui
from datetime import datetime
from functools import partial
from urllib.parse import quote
from openpyxl import load_workbook
import openpyxl
import csv


# Import a Browser to open the WhatsApp Web
import webbrowser


### Extras.py
def corrigir_numero(numero):
    if not numero:
        return ""

    numero = numero.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")

    if numero.startswith("0"):
        numero = numero[1:]
    numero = "+55" + numero
    return numero


def buscar_nome_telefone_novo(arquivo):

    tupla_nomes_telefones = []
    with open(arquivo, newline="") as csvfile:
        # Create a CSV reader object
        csvreader = csv.reader(csvfile, delimiter=",")

        # Process each row in the CSV file
        for row in csvreader:
            dia = row[12]
            nome = row[13]
            telefone = row[14]
            celular = row[15]

            try:
                int(dia)
            except:
                continue

            if not (telefone or celular):
                continue

            if celular:
                telefone = celular

            telefone = corrigir_numero(telefone)

            primeiro_nome = nome.split(" ")[0].capitalize()

            tupla_nomes_telefones.append((dia, nome, primeiro_nome, telefone))
    return tupla_nomes_telefones


def buscar_nome_telefone(arquivo):
    wb = openpyxl.load_workbook(arquivo)
    ws = wb.active
    tupla_nomes_telefones = []
    for row in ws.iter_rows(values_only=True):
        dia = row[0]
        nome = row[1]
        telefone = row[2]
        celular = row[3]

        try:
            int(row[0])
        except:
            continue

        if celular:
            telefone = celular

        if telefone:
            pass

        if not telefone:
            continue

        telefone = corrigir_numero(telefone)
        primeiro_nome = nome.split(" ")[0].capitalize()

        tupla_nomes_telefones.append((dia, nome, primeiro_nome, telefone))
    return tupla_nomes_telefones


### Extras.py


### FileSelector.py
class FileSelectorWindow(QtWidgets.QWidget):
    file_selected = QtCore.Signal(str)

    def __init__(self):
        super().__init__()
        ## Set Size
        self.setFixedSize(500, 100)
        ## Set Title
        self.setWindowTitle("SakaZap - Selecionar Arquivo")
        ## Set Layout
        self.layout = QtWidgets.QGridLayout(self)
        ## Create Widgets
        self.help_text = QtWidgets.QLabel("Selecione o arquivo que deseja acessar:")
        # Text Field
        self.text_field = QtWidgets.QLineEdit()
        self.text_field.setReadOnly(True)
        # Button
        self.button = QtWidgets.QPushButton("Abrir Arquivo")
        # Info Text
        self.info_text = QtWidgets.QLabel("Nenhum arquivo selecionado.")

        ## Add Widgets to Layout
        self.layout.addWidget(self.help_text, 0, 0)
        self.layout.addWidget(self.text_field, 1, 0)
        self.layout.addWidget(self.button, 1, 1)
        self.layout.addWidget(self.info_text, 2, 0)

        self.button.clicked.connect(self.open_file)

    def open_file(self):
        file_name, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, "Selecione o arquivo", "", "Arquivos CSV (*.csv)"
        )

        if file_name.endswith(".csv"):
            self.text_field.setText(file_name)
            self.info_text.setText(f"Arquivo Carregado com Sucesso!")
            self.file_selected.emit(file_name)
        else:
            self.info_text.setText("Por Favor Selecione um Arquivo CSV (.xlsx)")


### FileSelector.py


### UserTab.py
class UsersTab(QtWidgets.QWidget):
    def __init__(self, arquivo, message):
        super().__init__()
        tupla_nomes_telefones = buscar_nome_telefone_novo(arquivo)
        self.mensagem = message
        self.users_checkboxes = []

        main_layout = QtWidgets.QVBoxLayout(self)

        self.text = QtWidgets.QLabel(
            f"Aniversariantes do Dia de Hoje: {datetime.today().day}/{datetime.today().month}",
            alignment=QtCore.Qt.AlignLeft,
        )
        main_layout.addWidget(self.text)

        users_widget = QtWidgets.QWidget()
        users_layout = QtWidgets.QVBoxLayout(users_widget)

        self._gerar_campos_usuarios(tupla_nomes_telefones, users_layout)

        scroll_area = QtWidgets.QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setWidget(users_widget)
        main_layout.addWidget(scroll_area)

        report_button = QtWidgets.QPushButton("Gerar Relatório")
        report_button.clicked.connect(self.gerar_relatorio)
        main_layout.addWidget(report_button)

    def _gerar_campos_usuarios(self, users, layout):
        for dia, nome, primeiro_nome, telefone in users:
            if int(dia) != datetime.today().day:
                continue

            user_layout = QtWidgets.QHBoxLayout()
            user_layout.addWidget(QtWidgets.QLabel(nome))

            send_msg_button = QtWidgets.QPushButton("Enviar Mensagem")
            message_chkbox = QtWidgets.QCheckBox()
            message_chkbox.setDisabled(True)
            invalid_number_chkbox = QtWidgets.QCheckBox(
                "Número Inválido"
            )  # Nova Checkbox

            send_msg_button.clicked.connect(
                partial(
                    self.enviar_mensagem, telefone, nome, primeiro_nome, message_chkbox
                )
            )

            user_layout.addStretch()
            user_layout.addWidget(send_msg_button)
            user_layout.addWidget(message_chkbox)
            user_layout.addWidget(
                invalid_number_chkbox
            )  # Adicionar a nova checkbox ao layout
            layout.addLayout(user_layout)

            self.users_checkboxes.append(
                (nome, telefone, message_chkbox, invalid_number_chkbox)
            )

    def enviar_mensagem(self, telefone, nome, primeiro_nome, checkbox):
        texto = (
            self.mensagem.replace("NOMEPESSOA", primeiro_nome) if self.mensagem else ""
        )
        texto = texto.replace(
            "CUPOMPESSOA", f"{nome[:3]}{telefone[8:11]}{datetime.today().day}"
        )
        checkbox.setChecked(True)
        texto = quote(texto)
        url = f"https://web.whatsapp.com/send/?phone={telefone}&text={texto}"
        webbrowser.open(url)

    def gerar_relatorio(self):
        filename = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Salvar Relatório",
            f"Relatório_Aniversariantes_Dia_{datetime.today().strftime('%d_%m_%Y')}.xlsx",
            "Arquivos Excel (*.xslx)",
        )[0]

        if filename:
            workbook = load_workbook("template.xlsx")
            sheet = workbook.active

            row = 3
            for nome, telefone, msg_checkbox, invalid_chkbox in self.users_checkboxes:
                status = "SIM" if msg_checkbox.isChecked() else "NÃO"
                invalid_status = (
                    "Número inválido" if invalid_chkbox.isChecked() else None
                )
                dia = datetime.today().day
                sheet[f"B{row}"] = dia
                sheet[f"C{row}"] = nome
                sheet[f"D{row}"] = telefone
                sheet[f"E{row}"] = status
                sheet[f"F{row}"] = (
                    invalid_status  # Adicionar o status de número inválido
                )
                sheet[f"G{row}"] = f"{nome[:3]}{telefone[8:11]}{dia}"
                row += 1

            workbook.save(filename)


### UserTab.py


### MessageTab.py
class MessagesTab(QtWidgets.QWidget):
    def __init__(self):
        super().__init__()

        layout = QtWidgets.QVBoxLayout(self)

        self.text = QtWidgets.QLabel(
            "Editar Mensagem a Ser Enviada.", alignment=QtCore.Qt.AlignLeft
        )

        self.text_box = QtWidgets.QTextEdit()
        self.text_box.setStyleSheet("background-color: #d3d3d3; color: #000000;")
        self.text_box.setPlaceholderText("Digite sua mensagem aqui...")

        layout.addWidget(self.text)
        layout.addWidget(self.text_box)

        self.button = QtWidgets.QPushButton("Salvar Mensagem")
        layout.addWidget(self.button)

        self.button.clicked.connect(self.salvar_mensagem)

    def salvar_mensagem(self):
        with open("confs.cfg", "w", encoding="utf-8") as f:
            f.write(self.text_box.toPlainText())
        self.text_box.setText(self.text_box.toPlainText())
        self.text.setText("Mensagem Salva com Sucesso!")


### MessageTab.py


### Main.py
class Aplicativo(QtWidgets.QMainWindow):
    def update_message(self, message):
        # Update the message
        self.message = message

        # Update the UI to reflect the new message
        self.text_box.setText(self.message)

    def __init__(self, arquivo):
        super().__init__()
        self.setWindowTitle("SakaZap")
        self.arquivo = arquivo

        self.setFixedSize(600, 400)
        texto = self._inicializar_storage()

        self.tabWidget = QtWidgets.QTabWidget()
        self.setCentralWidget(self.tabWidget)

        self.messagesTab = MessagesTab()
        # Set the text in the message tab
        self.messagesTab.text_box.setText(texto)
        # Connect the clicked signal of the save button to the updateUsersTab function
        self.messagesTab.button.clicked.connect(self.updateUsersTab)

        self.usersTab = UsersTab(arquivo, texto)
        users = self.tabWidget.addTab(self.usersTab, "Clientes")
        messages = self.tabWidget.addTab(self.messagesTab, "Mensagem")

    def _inicializar_storage(self):
        try:
            f = open("confs.cfg", "r+", encoding="utf-8")
            return f.read()
        except FileNotFoundError:
            f = open("confs.cfg", "w+", encoding="utf-8")
            return ""

    def updateUsersTab(self):
        # Get the current message
        texto = self.messagesTab.text_box.toPlainText()

        # Remove the current UsersTab
        self.tabWidget.removeTab(0)

        # Create a new UsersTab with the updated message
        self.usersTab = UsersTab(self.arquivo, texto)

        # Add the new UsersTab to the QTabWidget
        self.tabWidget.insertTab(0, self.usersTab, "Clientes")


class MainApp(QtWidgets.QApplication):
    def __init__(self, sys_argv):
        super().__init__(sys_argv)
        self.file_selector = FileSelectorWindow()
        self.setWindowIcon(QtGui.QIcon("icon.png"))
        self.file_selector.file_selected.connect(self.load_main_app)
        self.file_selector.show()

    def load_main_app(self, file_name):
        try:
            self.file_selector.close()
            self.aplicativo = Aplicativo(file_name)
            self.aplicativo.show()
        except Exception as e:
            print(f"An error occurred: {e}")
            self.file_selector.show()


if __name__ == "__main__":
    app = MainApp(sys.argv)
    sys.exit(app.exec())
